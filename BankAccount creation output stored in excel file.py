# -*- coding: utf-8 -*-
"""
Created on Thu Aug  8 19:28:54 2024

@author: Naveena palanivelu
"""
import uuid
import openpyxl 

class Bank_Account:
    def createnewacc(self1):
        self1.Name=str(input("\n Enter the Customer Name: "))
        self1.Age=int(input("\n Enter the Customer age: "))
        self1.Gender=str(input("\n Enter the Customer Gender: "))
        self1.City=str(input("\n Enter the Customer City: "))
        self1.customerid =str(uuid.uuid4())
        self1.balance=0
        print("Welcome to the XXX Bank!!!!")
 
    def deposit(self1):
        amount=float(input("Enter amount to be Deposited: "))
        self1.balance += amount
        print("\n Amount Deposited:",amount)
 
    def withdraw(self1):
        amount = float(input("Enter amount to be Withdrawn: "))
        if self1.balance>=amount:
            self1.balance-=amount
            print("\n You Withdrew:", amount)
        else:
            print("\n Balance is not sufficient to withdraw")
 
    def display(self1):
        print("\n Customer id:",self1.customerid)
        print("\n Customer Name: ",self1.Name)
        print("\n Customer Age:",self1.Age)
        print("\n Customer Gender :",self1.Gender)
        print("\n Customer city :",self1.City)
        print("\n Account Balance",self1.balance)
        wb = openpyxl.Workbook() 
        sheet = wb.active 
        #print(x)
        c1 = sheet.cell(row=1, column=1)
        c1.value = "customerid"
        c2 = sheet.cell(row=1, column=2) 
        c2.value = "Name"
        c21 = sheet.cell(row=1, column=3) 
        c21.value = "Age"
        c22 = sheet.cell(row=1, column=4) 
        c22.value = "Gender"
        c23 = sheet.cell(row=1, column=5) 
        c23.value = "City"
        c24 = sheet.cell(row=1, column=6) 
        c24.value = "Balance"
        c3 = sheet.cell(row= 2, column=1)
        c3.value = self1.customerid
        c4 = sheet.cell(row=2, column=2)
        c4.value = self1.Name
        c4 = sheet.cell(row=2, column=3)
        c4.value = self1.Age
        c5 = sheet.cell(row=2, column=4)
        c5.value = self1.Gender
        c6 = sheet.cell(row=2, column=5)
        c6.value = self1.City
        c7 = sheet.cell(row=2, column=6)
        c7.value = self1.balance
        
        wb.save("output.xlsx")

  
# creating an object of class
"""i=2
for x in range (0,i):"""
s = Bank_Account()
s.createnewacc()
s.deposit()
s.withdraw()
s.display()
# Calling functions with that class object

